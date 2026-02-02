"""Base class for provincial ingesters.

This module provides the abstract base classes that all provincial ingesters
inherit from. It handles common functionality like:
- Multi-format file download with progress (XLSX, CSV, XML, JSON)
- Record hash computation for incremental sync
- Entity matching with fuzzy name matching
- PostgreSQL and Neo4j sync

## Adding a New Provincial Ingester

### For Non-Profit Organizations (004 legacy)

Inherit from `BaseProvincialIngester` for XLSX-only non-profit ingesters.

### For All Corporation Types (005+)

Inherit from `BaseProvincialCorpIngester` for multi-format support.

Example:
    ```python
    class QuebecCorporationIngester(BaseProvincialCorpIngester):
        @property
        def province(self) -> str:
            return "QC"

        @property
        def data_format(self) -> str:
            return "csv"

        def get_data_url(self) -> str:
            return "https://example.com/quebec-corps.csv"

        def get_csv_encoding(self) -> str:
            return "utf-8"

        def parse_record(self, row: tuple) -> ProvincialCorporationRecord | None:
            ...
    ```
"""

import csv
import json
import sys
from abc import ABC, abstractmethod
from collections.abc import AsyncIterator
from datetime import datetime
from io import BytesIO, StringIO
from typing import Any, Literal
from uuid import UUID, uuid4

from openpyxl import load_workbook
from rapidfuzz import fuzz
from sqlalchemy import text

from ..base import (
    BaseIngester,
    IngestionConfig,
    Neo4jHelper,
    PostgresHelper,
    download_with_progress,
)
from .models import (
    EntityMatchResult,
    ProvincialCorporationRecord,
    ProvincialCorpStatus,
    ProvincialCorpType,
    ProvincialNonProfitRecord,
)


class BaseProvincialIngester(BaseIngester[ProvincialNonProfitRecord], ABC):
    """Abstract base class for provincial non-profit ingesters.

    Provides common functionality for all provincial data sources:
    - XLSX download and parsing
    - Record deduplication via hash
    - Entity matching (exact and fuzzy)
    - PostgreSQL and Neo4j sync

    Subclasses must implement:
    - province: Province code (e.g., "AB", "ON")
    - get_data_url(): URL to download data
    - parse_record(row): Parse a row into ProvincialNonProfitRecord
    """

    # Fuzzy matching thresholds
    FUZZY_MATCH_THRESHOLD = 0.85  # Minimum score for fuzzy match
    AUTO_LINK_THRESHOLD = 0.95   # Auto-link without review above this

    def __init__(self):
        """Initialize the provincial ingester."""
        super().__init__(f"{self.province.lower()}-nonprofits")
        self._neo4j = Neo4jHelper(self.logger)
        self._postgres = PostgresHelper(self.logger)
        self._record_hashes: dict[str, str] = {}  # name -> hash mapping

    @property
    @abstractmethod
    def province(self) -> str:
        """Return the province code (e.g., 'AB', 'ON', 'BC').

        Must be a valid 2-letter Canadian province/territory code.
        """
        ...

    @abstractmethod
    def get_data_url(self) -> str:
        """Return the URL to download the data file.

        Returns:
            URL to the XLSX/CSV data file
        """
        ...

    @abstractmethod
    def parse_record(self, row: tuple) -> ProvincialNonProfitRecord | None:
        """Parse a single row from the data file.

        Args:
            row: Tuple of values from the spreadsheet row

        Returns:
            ProvincialNonProfitRecord if valid, None to skip the row
        """
        ...

    def get_expected_columns(self) -> list[str] | None:
        """Return expected column names for validation.

        Override this to validate columns on startup. If the actual
        columns don't match, ingestion will fail with a descriptive error.

        Returns:
            List of expected column names, or None to skip validation
        """
        return None

    def get_header_row_index(self) -> int:
        """Return the 0-based index of the header row in the XLSX file.

        Override this if the data file has empty rows before the header.
        Default is 0 (first row is the header).

        Returns:
            0-based row index where column headers are located
        """
        return 0

    async def download_xlsx(self) -> bytes:
        """Download the XLSX file with progress bar.

        Returns:
            Raw bytes of the downloaded file
        """
        url = self.get_data_url()
        self.logger.info(f"Downloading data from {url}")
        return await download_with_progress(
            url,
            desc=f"Downloading {self.province} non-profit data",
        )

    def parse_xlsx(self, data: bytes) -> list[tuple]:
        """Parse XLSX data into rows.

        Args:
            data: Raw XLSX bytes

        Returns:
            List of row tuples (excluding header and pre-header rows)
        """
        # Load workbook in read-only mode for memory efficiency
        wb = load_workbook(filename=BytesIO(data), read_only=True, data_only=True)
        ws = wb.active

        rows = []
        header_row_index = self.get_header_row_index()

        for i, row in enumerate(ws.iter_rows(values_only=True)):
            # Skip rows before header
            if i < header_row_index:
                continue
            if i == header_row_index:
                # Validate columns if expected columns are defined
                expected = self.get_expected_columns()
                if expected:
                    actual = [str(c).strip() if c else "" for c in row]
                    missing = set(expected) - set(actual)
                    if missing:
                        raise ValueError(
                            f"Missing expected columns: {missing}. "
                            f"Actual columns: {actual}"
                        )
                continue
            rows.append(row)

        wb.close()
        return rows

    async def fetch_records(
        self, config: IngestionConfig
    ) -> AsyncIterator[ProvincialNonProfitRecord]:
        """Fetch and parse records from the provincial data source.

        Downloads the XLSX file, parses it, and yields valid records.
        Supports incremental sync by comparing record hashes.

        Args:
            config: Ingestion configuration

        Yields:
            ProvincialNonProfitRecord for each valid row
        """
        # Download the data
        data = await self.download_xlsx()

        # Parse XLSX
        print("Parsing XLSX file...", file=sys.stderr, end=" ")
        rows = self.parse_xlsx(data)
        print(f"found {len(rows):,} records", file=sys.stderr)

        # Load existing hashes for incremental sync
        if config.incremental:
            await self._load_existing_hashes()

        # Process rows
        processed = 0
        for row in rows:
            try:
                record = self.parse_record(row)
                if record is None:
                    continue

                # Check if changed (incremental sync)
                if config.incremental:
                    new_hash = record.compute_record_hash()
                    existing_hash = self._record_hashes.get(record.name)
                    if existing_hash == new_hash:
                        # No change, skip
                        continue

                yield record
                processed += 1

                # Check limit
                if config.limit and processed >= config.limit:
                    break

            except Exception as e:
                self.logger.warning(f"Failed to parse row: {e}")
                continue

    async def _load_existing_hashes(self) -> None:
        """Load existing record hashes from database for incremental sync."""
        from ...db import get_db_session

        self._record_hashes = {}

        try:
            async with get_db_session() as db:
                result = await db.execute(
                    text("""
                        SELECT name, metadata->>'record_hash' as hash
                        FROM entities
                        WHERE provincial_registry_id LIKE :province_prefix
                        AND metadata->>'record_hash' IS NOT NULL
                    """),
                    {"province_prefix": f"{self.province}:%"},
                )
                for row in result.fetchall():
                    if row.hash:
                        self._record_hashes[row.name] = row.hash

            self.logger.info(
                f"Loaded {len(self._record_hashes)} existing record hashes"
            )
        except Exception as e:
            self.logger.warning(f"Failed to load existing hashes: {e}")

    async def process_record(
        self, record: ProvincialNonProfitRecord
    ) -> dict[str, Any]:
        """Process a single provincial non-profit record.

        Stores the record in PostgreSQL and syncs to Neo4j.
        Performs entity matching to link with existing entities.

        Args:
            record: The record to process

        Returns:
            Processing result with status and details
        """
        from ...db import get_db_session, get_neo4j_session

        # Try to match with existing entity
        match_result = await self.match_existing_entity(record)

        async with get_db_session() as db:
            if match_result.is_match and match_result.is_auto_linkable:
                # Link to existing entity
                entity_id = match_result.matched_entity_id
                is_new = False

                # Update existing entity with provincial data
                await db.execute(
                    text("""
                        UPDATE entities SET
                            provincial_registry_id = :registry_id,
                            metadata = COALESCE(metadata, '{}'::jsonb) || CAST(:metadata AS jsonb),
                            updated_at = :updated_at
                        WHERE id = :id
                    """),
                    {
                        "id": entity_id,
                        "registry_id": record.provincial_registry_id,
                        "metadata": json.dumps({
                            "provincial_org_type": record.org_type_parsed.value,
                            "provincial_status": record.status_parsed.value,
                            "registration_date": record.registration_date.isoformat() if record.registration_date else None,
                            "city": record.city,
                            "postal_code": record.postal_code,
                            "source_url": record.source_url,
                            "record_hash": record.compute_record_hash(),
                            "last_synced": datetime.utcnow().isoformat(),
                        }),
                        "updated_at": datetime.utcnow(),
                    },
                )
                result_type = "updated"

            elif match_result.requires_review:
                # Flag for manual review - create new entity but mark it
                entity_id, is_new = await self._create_provincial_entity(db, record)
                await self._log_match_for_review(db, record, match_result)
                result_type = "created"

            else:
                # No match - create new entity
                entity_id, is_new = await self._create_provincial_entity(db, record)
                result_type = "created" if is_new else "duplicate"

        # Sync to Neo4j
        try:
            async with get_neo4j_session() as session:
                await self._neo4j.merge_organization(
                    session,
                    id=str(entity_id),
                    name=record.name,
                    org_type="nonprofit",
                    external_ids={
                        "provincial_registry_id": record.provincial_registry_id,
                    },
                    properties={
                        "jurisdiction": f"CA-{record.province}",
                        "provincial_org_type": record.org_type_parsed.value,
                        "provincial_status": record.status_parsed.value,
                        "city": record.city,
                        "postal_code": record.postal_code,
                        "registration_date": record.registration_date.isoformat() if record.registration_date else None,
                    },
                )
        except Exception as e:
            self.logger.warning(f"Neo4j sync failed for {record.name}: {e}")

        return {
            result_type: True,
            "entity_id": str(entity_id),
            "match_result": match_result.model_dump() if match_result.is_match else None,
        }

    async def _create_provincial_entity(
        self,
        db,
        record: ProvincialNonProfitRecord,
    ) -> tuple[UUID, bool]:
        """Create a new entity from provincial record.

        Args:
            db: Database session
            record: Provincial record

        Returns:
            Tuple of (entity_id, is_new)
        """
        # Check if entity already exists by provincial_registry_id
        result = await db.execute(
            text("""
                SELECT id FROM entities
                WHERE provincial_registry_id = :registry_id
            """),
            {"registry_id": record.provincial_registry_id},
        )
        existing = result.fetchone()

        now = datetime.utcnow()
        metadata = {
            "provincial_org_type": record.org_type_parsed.value,
            "provincial_status": record.status_parsed.value,
            "registration_date": record.registration_date.isoformat() if record.registration_date else None,
            "city": record.city,
            "postal_code": record.postal_code,
            "source_url": record.source_url,
            "record_hash": record.compute_record_hash(),
            "last_synced": now.isoformat(),
        }

        if existing:
            # Update existing
            await db.execute(
                text("""
                    UPDATE entities SET
                        metadata = COALESCE(metadata, '{}'::jsonb) || CAST(:metadata AS jsonb),
                        updated_at = :updated_at
                    WHERE id = :id
                """),
                {
                    "id": existing.id,
                    "metadata": json.dumps(metadata),
                    "updated_at": now,
                },
            )
            return existing.id, False
        else:
            # Create new
            new_id = uuid4()
            await db.execute(
                text("""
                    INSERT INTO entities (
                        id, name, entity_type, provincial_registry_id,
                        external_ids, metadata, created_at, updated_at
                    ) VALUES (
                        :id, :name, :entity_type, :registry_id,
                        CAST(:external_ids AS jsonb), CAST(:metadata AS jsonb),
                        :created_at, :updated_at
                    )
                """),
                {
                    "id": new_id,
                    "name": record.name,
                    "entity_type": "organization",
                    "registry_id": record.provincial_registry_id,
                    "external_ids": json.dumps({
                        "provincial_registry": record.province,
                        f"{record.province.lower()}_registry_id": record.provincial_registry_id,
                    }),
                    "metadata": json.dumps(metadata),
                    "created_at": now,
                    "updated_at": now,
                },
            )
            return new_id, True

    async def match_existing_entity(
        self, record: ProvincialNonProfitRecord
    ) -> EntityMatchResult:
        """Match a provincial record against existing entities.

        Performs exact name matching first, then fuzzy matching.

        Args:
            record: Provincial record to match

        Returns:
            EntityMatchResult with match details
        """
        from ...db import get_db_session

        async with get_db_session() as db:
            # Try exact match first
            result = await db.execute(
                text("""
                    SELECT id, name FROM entities
                    WHERE LOWER(name) = LOWER(:name)
                    AND entity_type = 'organization'
                    AND provincial_registry_id IS NULL
                    LIMIT 1
                """),
                {"name": record.name},
            )
            exact_match = result.fetchone()

            if exact_match:
                return EntityMatchResult(
                    provincial_record_name=record.name,
                    matched_entity_id=exact_match.id,
                    matched_entity_name=exact_match.name,
                    match_score=1.0,
                    match_method="exact",
                    requires_review=False,
                )

            # Try fuzzy match
            result = await db.execute(
                text("""
                    SELECT id, name FROM entities
                    WHERE entity_type = 'organization'
                    AND provincial_registry_id IS NULL
                    AND similarity(LOWER(name), LOWER(:name)) > 0.3
                    ORDER BY similarity(LOWER(name), LOWER(:name)) DESC
                    LIMIT 5
                """),
                {"name": record.name},
            )
            candidates = result.fetchall()

            best_match = None
            best_score = 0.0

            for candidate in candidates:
                # Use token_sort_ratio for better handling of word order
                score = fuzz.token_sort_ratio(
                    record.name.lower(),
                    candidate.name.lower(),
                ) / 100.0

                if score > best_score:
                    best_score = score
                    best_match = candidate

            if best_match and best_score >= self.FUZZY_MATCH_THRESHOLD:
                return EntityMatchResult(
                    provincial_record_name=record.name,
                    matched_entity_id=best_match.id,
                    matched_entity_name=best_match.name,
                    match_score=best_score,
                    match_method="fuzzy",
                    requires_review=best_score < self.AUTO_LINK_THRESHOLD,
                )

        # No match found
        return EntityMatchResult(
            provincial_record_name=record.name,
            match_score=0.0,
            match_method="none",
        )

    async def _log_match_for_review(
        self,
        _db,  # Reserved for future use (storing in review table)
        record: ProvincialNonProfitRecord,
        match_result: EntityMatchResult,
    ) -> None:
        """Log a fuzzy match that requires manual review.

        Args:
            _db: Database session (reserved for future use)
            record: Provincial record
            match_result: Match result with details
        """
        # TODO: Store in evidence table or a dedicated review table
        # For now, just log to ingestion metadata
        self.logger.info(
            f"Match requires review: '{record.name}' -> "
            f"'{match_result.matched_entity_name}' "
            f"(score: {match_result.match_score:.2%})"
        )

    async def get_last_sync_time(self) -> datetime | None:
        """Get the timestamp of the last successful sync."""
        from ...db import get_db_session

        try:
            async with get_db_session() as db:
                result = await db.execute(
                    text("""
                        SELECT completed_at FROM ingestion_runs
                        WHERE source = :source
                        AND status IN ('completed', 'partial')
                        ORDER BY completed_at DESC
                        LIMIT 1
                    """),
                    {"source": self.source_name},
                )
                row = result.fetchone()
                return row.completed_at if row else None
        except Exception:
            return None

    async def save_sync_time(self, timestamp: datetime) -> None:
        """Save the timestamp of a successful sync.

        Note: This is handled automatically by the base ingester's run() method.
        """
        pass


# =============================================================================
# Base class for provincial corporation ingesters (005+)
# =============================================================================


class BaseProvincialCorpIngester(BaseIngester[ProvincialCorporationRecord], ABC):
    """Abstract base class for provincial corporation ingesters.

    Provides multi-format support (XLSX, CSV, XML, JSON) for all corporation types.
    Supports for-profit, non-profit, cooperative, and other corporation types.

    Subclasses must implement:
    - province: Province code (e.g., "AB", "ON")
    - data_format: Data format ("xlsx", "csv", "xml", "json")
    - get_data_url(): URL to download data
    - parse_record(row): Parse a row into ProvincialCorporationRecord
    """

    # Fuzzy matching thresholds
    FUZZY_MATCH_THRESHOLD = 0.85  # Minimum score for fuzzy match
    AUTO_LINK_THRESHOLD = 0.95   # Auto-link without review above this

    def __init__(self, source_suffix: str = "corps"):
        """Initialize the provincial corporation ingester.

        Args:
            source_suffix: Suffix for source name (default: "corps")
        """
        super().__init__(f"{self.province.lower()}-{source_suffix}")
        self._neo4j = Neo4jHelper(self.logger)
        self._postgres = PostgresHelper(self.logger)
        self._record_hashes: dict[str, str] = {}

    @property
    @abstractmethod
    def province(self) -> str:
        """Return the province code (e.g., 'AB', 'ON', 'QC')."""
        ...

    @property
    @abstractmethod
    def data_format(self) -> Literal["xlsx", "csv", "xml", "json"]:
        """Return the data format this ingester handles."""
        ...

    @abstractmethod
    def get_data_url(self) -> str:
        """Return the URL to download the data file."""
        ...

    @abstractmethod
    def parse_record(self, row: tuple | dict) -> ProvincialCorporationRecord | None:
        """Parse a single row/record from the data file.

        Args:
            row: Tuple (for XLSX/CSV) or dict (for JSON/XML) of values

        Returns:
            ProvincialCorporationRecord if valid, None to skip
        """
        ...

    def get_expected_columns(self) -> list[str] | None:
        """Return expected column names for validation (XLSX/CSV only)."""
        return None

    def get_header_row_index(self) -> int:
        """Return 0-based header row index (XLSX only). Default: 0."""
        return 0

    def get_csv_encoding(self) -> str:
        """Return CSV file encoding. Default: utf-8."""
        return "utf-8"

    def get_csv_delimiter(self) -> str:
        """Return CSV delimiter character. Default: comma."""
        return ","

    async def download_data(self) -> bytes:
        """Download the data file with progress bar."""
        url = self.get_data_url()
        self.logger.info(f"Downloading {self.province} data from {url}")
        return await download_with_progress(
            url,
            desc=f"Downloading {self.province} corporation data",
        )

    def parse_xlsx(self, data: bytes) -> list[tuple]:
        """Parse XLSX data into rows."""
        wb = load_workbook(filename=BytesIO(data), read_only=True, data_only=True)
        ws = wb.active

        rows = []
        header_row_index = self.get_header_row_index()

        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i < header_row_index:
                continue
            if i == header_row_index:
                expected = self.get_expected_columns()
                if expected:
                    actual = [str(c).strip() if c else "" for c in row]
                    missing = set(expected) - set(actual)
                    if missing:
                        raise ValueError(
                            f"Missing expected columns: {missing}. Actual: {actual}"
                        )
                continue
            rows.append(row)

        wb.close()
        return rows

    def parse_csv(self, data: bytes) -> list[tuple]:
        """Parse CSV data into rows.

        Args:
            data: Raw CSV bytes

        Returns:
            List of row tuples (excluding header)
        """
        encoding = self.get_csv_encoding()
        delimiter = self.get_csv_delimiter()

        # Decode bytes to string
        text_data = data.decode(encoding)
        reader = csv.reader(StringIO(text_data), delimiter=delimiter)

        rows = []
        header = None

        for i, row in enumerate(reader):
            if i == 0:
                header = row
                # Validate columns if expected
                expected = self.get_expected_columns()
                if expected:
                    actual = [str(c).strip() for c in header]
                    missing = set(expected) - set(actual)
                    if missing:
                        raise ValueError(
                            f"Missing expected columns: {missing}. Actual: {actual}"
                        )
                continue
            rows.append(tuple(row))

        return rows

    def parse_xml(self, data: bytes) -> list[dict]:
        """Parse XML data into records.

        Override this method for province-specific XML structures.

        Args:
            data: Raw XML bytes

        Returns:
            List of record dictionaries
        """
        try:
            from lxml import etree
        except ImportError:
            raise ImportError("lxml is required for XML parsing. Install it with: pip install lxml")

        tree = etree.fromstring(data)
        records = []

        # Generic XML parsing - override for specific structures
        for element in tree.iter():
            if element.text and element.text.strip():
                records.append({element.tag: element.text.strip()})

        return records

    def parse_json(self, data: bytes) -> list[dict]:
        """Parse JSON data into records.

        Args:
            data: Raw JSON bytes

        Returns:
            List of record dictionaries
        """
        parsed = json.loads(data.decode("utf-8"))

        # Handle common JSON structures
        if isinstance(parsed, list):
            return parsed
        elif isinstance(parsed, dict):
            # Look for common data keys
            for key in ["data", "records", "results", "items", "corporations"]:
                if key in parsed and isinstance(parsed[key], list):
                    return parsed[key]
            # Return single record as list
            return [parsed]

        return []

    async def fetch_records(
        self, config: IngestionConfig
    ) -> AsyncIterator[ProvincialCorporationRecord]:
        """Fetch and parse records from the provincial data source.

        Downloads the data file, parses based on format, and yields valid records.
        Supports incremental sync by comparing record hashes.
        """
        # Download the data
        data = await self.download_data()

        # Parse based on format
        format_name = self.data_format.upper()
        print(f"Parsing {format_name} file...", file=sys.stderr, end=" ")

        if self.data_format == "xlsx":
            rows = self.parse_xlsx(data)
        elif self.data_format == "csv":
            rows = self.parse_csv(data)
        elif self.data_format == "xml":
            rows = self.parse_xml(data)
        elif self.data_format == "json":
            rows = self.parse_json(data)
        else:
            raise ValueError(f"Unsupported data format: {self.data_format}")

        print(f"found {len(rows):,} records", file=sys.stderr)

        # Load existing hashes for incremental sync
        if config.incremental:
            await self._load_existing_hashes()

        # Process rows
        processed = 0
        for row in rows:
            try:
                record = self.parse_record(row)
                if record is None:
                    continue

                # Check if changed (incremental sync)
                if config.incremental:
                    new_hash = record.compute_record_hash()
                    existing_hash = self._record_hashes.get(record.name)
                    if existing_hash == new_hash:
                        continue

                yield record
                processed += 1

                if config.limit and processed >= config.limit:
                    break

            except Exception as e:
                self.logger.warning(f"Failed to parse row: {e}")
                continue

    async def _load_existing_hashes(self) -> None:
        """Load existing record hashes from database for incremental sync."""
        from ...db import get_db_session

        self._record_hashes = {}

        try:
            async with get_db_session() as db:
                result = await db.execute(
                    text("""
                        SELECT name, metadata->>'record_hash' as hash
                        FROM entities
                        WHERE provincial_registry_id LIKE :province_prefix
                        AND metadata->>'record_hash' IS NOT NULL
                    """),
                    {"province_prefix": f"{self.province}:%"},
                )
                for row in result.fetchall():
                    if row.hash:
                        self._record_hashes[row.name] = row.hash

            self.logger.info(
                f"Loaded {len(self._record_hashes)} existing record hashes"
            )
        except Exception as e:
            self.logger.warning(f"Failed to load existing hashes: {e}")

    async def process_record(
        self, record: ProvincialCorporationRecord
    ) -> dict[str, Any]:
        """Process a single provincial corporation record.

        Stores the record in PostgreSQL and syncs to Neo4j.
        """
        from ...db import get_db_session, get_neo4j_session

        # Try to match with existing entity
        match_result = await self.match_existing_entity(record)

        async with get_db_session() as db:
            if match_result.is_match and match_result.is_auto_linkable:
                entity_id = match_result.matched_entity_id
                await self._update_entity_with_provincial_data(db, entity_id, record)
                result_type = "updated"
            elif match_result.requires_review:
                entity_id, is_new = await self._create_corp_entity(db, record)
                await self._log_match_for_review(db, record, match_result)
                result_type = "created"
            else:
                entity_id, is_new = await self._create_corp_entity(db, record)
                result_type = "created" if is_new else "duplicate"

        # Sync to Neo4j
        try:
            async with get_neo4j_session() as session:
                await self._neo4j.merge_organization(
                    session,
                    id=str(entity_id),
                    name=record.name,
                    org_type=record.corp_type_parsed.value,
                    external_ids={
                        "provincial_registry_id": record.provincial_registry_id,
                        "business_number": record.business_number,
                    },
                    properties={
                        "jurisdiction": f"CA-{record.jurisdiction}",
                        "provincial_corp_type": record.corp_type_parsed.value,
                        "provincial_status": record.status_parsed.value,
                        "incorporation_date": record.incorporation_date.isoformat() if record.incorporation_date else None,
                        "name_french": record.name_french,
                    },
                )
        except Exception as e:
            self.logger.warning(f"Neo4j sync failed for {record.name}: {e}")

        return {
            result_type: True,
            "entity_id": str(entity_id),
            "match_result": match_result.model_dump() if match_result.is_match else None,
        }

    async def _update_entity_with_provincial_data(
        self, db, entity_id: UUID, record: ProvincialCorporationRecord
    ) -> None:
        """Update existing entity with provincial data."""
        now = datetime.utcnow()
        metadata = self._build_metadata(record, now)

        await db.execute(
            text("""
                UPDATE entities SET
                    provincial_registry_id = :registry_id,
                    metadata = COALESCE(metadata, '{}'::jsonb) || CAST(:metadata AS jsonb),
                    updated_at = :updated_at
                WHERE id = :id
            """),
            {
                "id": entity_id,
                "registry_id": record.provincial_registry_id,
                "metadata": json.dumps(metadata),
                "updated_at": now,
            },
        )

    async def _create_corp_entity(
        self, db, record: ProvincialCorporationRecord
    ) -> tuple[UUID, bool]:
        """Create a new entity from provincial corporation record."""
        result = await db.execute(
            text("SELECT id FROM entities WHERE provincial_registry_id = :registry_id"),
            {"registry_id": record.provincial_registry_id},
        )
        existing = result.fetchone()

        now = datetime.utcnow()
        metadata = self._build_metadata(record, now)

        if existing:
            await db.execute(
                text("""
                    UPDATE entities SET
                        metadata = COALESCE(metadata, '{}'::jsonb) || CAST(:metadata AS jsonb),
                        updated_at = :updated_at
                    WHERE id = :id
                """),
                {"id": existing.id, "metadata": json.dumps(metadata), "updated_at": now},
            )
            return existing.id, False
        else:
            new_id = uuid4()
            external_ids = {
                "provincial_registry": record.jurisdiction,
                f"{record.jurisdiction.lower()}_corp_number": record.registration_number,
            }
            if record.business_number:
                external_ids["business_number"] = record.business_number

            await db.execute(
                text("""
                    INSERT INTO entities (
                        id, name, entity_type, provincial_registry_id,
                        external_ids, metadata, created_at, updated_at
                    ) VALUES (
                        :id, :name, :entity_type, :registry_id,
                        CAST(:external_ids AS jsonb), CAST(:metadata AS jsonb),
                        :created_at, :updated_at
                    )
                """),
                {
                    "id": new_id,
                    "name": record.name,
                    "entity_type": "organization",
                    "registry_id": record.provincial_registry_id,
                    "external_ids": json.dumps(external_ids),
                    "metadata": json.dumps(metadata),
                    "created_at": now,
                    "updated_at": now,
                },
            )
            return new_id, True

    def _build_metadata(
        self, record: ProvincialCorporationRecord, now: datetime
    ) -> dict:
        """Build metadata dict for entity storage."""
        addr = record.registered_address
        return {
            "provincial_corp_type": record.corp_type_parsed.value,
            "provincial_corp_type_raw": record.corp_type_raw,
            "provincial_status": record.status_parsed.value,
            "provincial_status_raw": record.status_raw,
            "incorporation_date": record.incorporation_date.isoformat() if record.incorporation_date else None,
            "jurisdiction": record.jurisdiction,
            "name_french": record.name_french,
            "registered_address": {
                "street": addr.street_address if addr else None,
                "city": addr.city if addr else None,
                "province": addr.province if addr else None,
                "postal_code": addr.postal_code if addr else None,
            } if addr else None,
            "source_url": record.source_url,
            "record_hash": record.compute_record_hash(),
            "last_synced": now.isoformat(),
        }

    async def match_existing_entity(
        self, record: ProvincialCorporationRecord
    ) -> EntityMatchResult:
        """Match a provincial record against existing entities.

        Uses business number, exact name, and fuzzy name matching.
        """
        from ...db import get_db_session

        async with get_db_session() as db:
            # Try business number match first (highest confidence)
            if record.business_number:
                result = await db.execute(
                    text("""
                        SELECT id, name FROM entities
                        WHERE external_ids->>'business_number' = :bn
                        LIMIT 1
                    """),
                    {"bn": record.business_number},
                )
                bn_match = result.fetchone()
                if bn_match:
                    return EntityMatchResult(
                        provincial_record_name=record.name,
                        matched_entity_id=bn_match.id,
                        matched_entity_name=bn_match.name,
                        match_score=1.0,
                        match_method="business_number",
                        requires_review=False,
                    )

            # Try exact name match
            result = await db.execute(
                text("""
                    SELECT id, name FROM entities
                    WHERE LOWER(name) = LOWER(:name)
                    AND entity_type = 'organization'
                    AND provincial_registry_id IS NULL
                    LIMIT 1
                """),
                {"name": record.name},
            )
            exact_match = result.fetchone()

            if exact_match:
                return EntityMatchResult(
                    provincial_record_name=record.name,
                    matched_entity_id=exact_match.id,
                    matched_entity_name=exact_match.name,
                    match_score=1.0,
                    match_method="exact",
                    requires_review=False,
                )

            # Try fuzzy match
            result = await db.execute(
                text("""
                    SELECT id, name FROM entities
                    WHERE entity_type = 'organization'
                    AND provincial_registry_id IS NULL
                    AND similarity(LOWER(name), LOWER(:name)) > 0.3
                    ORDER BY similarity(LOWER(name), LOWER(:name)) DESC
                    LIMIT 5
                """),
                {"name": record.name},
            )
            candidates = result.fetchall()

            best_match = None
            best_score = 0.0

            for candidate in candidates:
                score = fuzz.token_sort_ratio(
                    record.name.lower(), candidate.name.lower()
                ) / 100.0

                if score > best_score:
                    best_score = score
                    best_match = candidate

            if best_match and best_score >= self.FUZZY_MATCH_THRESHOLD:
                return EntityMatchResult(
                    provincial_record_name=record.name,
                    matched_entity_id=best_match.id,
                    matched_entity_name=best_match.name,
                    match_score=best_score,
                    match_method="fuzzy",
                    requires_review=best_score < self.AUTO_LINK_THRESHOLD,
                )

        return EntityMatchResult(
            provincial_record_name=record.name,
            match_score=0.0,
            match_method="none",
        )

    async def _log_match_for_review(
        self, _db, record: ProvincialCorporationRecord, match_result: EntityMatchResult
    ) -> None:
        """Log a fuzzy match that requires manual review."""
        self.logger.info(
            f"Match requires review: '{record.name}' -> "
            f"'{match_result.matched_entity_name}' "
            f"(score: {match_result.match_score:.2%})"
        )

    async def get_last_sync_time(self) -> datetime | None:
        """Get the timestamp of the last successful sync."""
        from ...db import get_db_session

        try:
            async with get_db_session() as db:
                result = await db.execute(
                    text("""
                        SELECT completed_at FROM ingestion_runs
                        WHERE source = :source
                        AND status IN ('completed', 'partial')
                        ORDER BY completed_at DESC
                        LIMIT 1
                    """),
                    {"source": self.source_name},
                )
                row = result.fetchone()
                return row.completed_at if row else None
        except Exception:
            return None

    async def save_sync_time(self, timestamp: datetime) -> None:
        """Save the timestamp of a successful sync."""
        pass
